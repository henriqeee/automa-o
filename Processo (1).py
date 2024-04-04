import openpyxl
import win32com.client
import time
import psutil
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.alignment import Alignment
from openpyxl import load_workbook
import schedule
import win32api
import os
sap_gui_auto = win32com.client.GetObject("SAPGUI")
application = sap_gui_auto.GetScriptingEngine
connection = application.Children(0)
session = connection.Children(0)



dt = datetime.now()
formatada = dt.strftime('%d.%m.20%y')
hora_formatada = dt.strftime('%H:%M:%S')
data = dt - timedelta(days=5)
data_5 = data.strftime('%d.%m.%Y')
time.sleep(2)


try:
    session.findById("wnd[0]/tbar[0]/okcd").text = "/n/scwm/mon"
    session.findById("wnd[0]").sendVKey (0)

    if session.ActiveWindow.Name == "wnd[1]":
        session.findById("wnd[1]/usr/ctxtP_LGNUM").text = "ibi"
        session.findById("wnd[1]/usr/ctxtP_MONIT").text = "z001"
        session.findById("wnd[1]/usr/ctxtP_MONIT").setFocus
        session.findById("wnd[1]/usr/ctxtP_MONIT").caretPosition = 4
        session.findById("wnd[1]").sendVKey (8)

    
    session.findById("wnd[0]/usr/shell/shellcont[0]/shell").selectedNode = "Z000000115" 
    session.findById("wnd[0]/usr/shell/shellcont[0]/shell").doubleClickNode ("Z000000115")
    session.findById("wnd[1]/usr/ctxtS_DOCTY-LOW").text = "zpat"
    session.findById("wnd[1]/usr/ctxtP_CDATFR").text = data_5
    session.findById("wnd[1]/usr/ctxtP_CDATTO").text = formatada
    session.findById("wnd[1]/usr/ctxtP_CTIMTO").text = hora_formatada
    session.findById("wnd[1]/usr/ctxtP_CTIMTO").setFocus
    session.findById("wnd[1]/usr/ctxtP_CTIMTO").caretPosition = 8
    session.findById("wnd[1]").sendVKey (8)
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").setCurrentCell(-1,"")
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").selectAll()
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").pressToolbarButton ("Z000000116")
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[1]/shell").pressToolbarContextButton ("&MB_VARIANT")
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[1]/shell").selectContextMenuItem ("&LOAD")
    session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cmbG51_SCREEN-USPEC_LBOX").key = "X"
    session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").clickCurrentCell()
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[1]/shell").pressToolbarContextButton ("&MB_EXPORT")
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[1]/shell").selectContextMenuItem ("&XXL")
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = r"C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels"
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "critico.XLSX"
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 12
    session.findById("wnd[1]/tbar[0]/btn[11]").press()
    time.sleep(10)
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").pressToolbarContextButton("METHODS")
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").selectContextMenuItem("@M00001")
    session.findById("wnd[1]").close()
    time.sleep(25)
    excel = win32com.client.Dispatch('Excel.Application')
    time.sleep(10)

    excel.Quit()


    time.sleep(25)
    caminho_arquivo_excel = r'C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels\critico.XLSX'

    df = pd.read_excel(caminho_arquivo_excel)
    nome_da_coluna = 'Produto'
    df[nome_da_coluna].to_clipboard(index=False)




    session.findById("wnd[0]").maximize
    session.findById("wnd[0]/usr/shell/shellcont[0]/shell").selectedNode = "N000000137"
    session.findById("wnd[0]/usr/shell/shellcont[0]/shell").doubleClickNode("N000000137")
    session.findById("wnd[1]/usr/btn%_S_MATNR_%_APP_%-VALU_PUSH").press()
    session.findById("wnd[2]/tbar[0]/btn[24]").press()
    session.findById("wnd[2]/tbar[0]/btn[8]").press()
    session.findById("wnd[1]/tbar[0]/btn[8]").press()
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").pressToolbarContextButton ("&MB_VARIANT")
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").selectContextMenuItem ("&LOAD")
    session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cmbG51_SCREEN-USPEC_LBOX").key = "X"
    session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").currentCellRow = 0
    session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").selectedRows = "0"
    session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").clickCurrentCell()
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").pressToolbarContextButton ("&MB_EXPORT")
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").selectContextMenuItem ("&XXL")
    session.findById("wnd[1]/usr/ctxtDY_PATH").text = r"C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels"
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "estoque_critico.XLSX"
    session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 1
    session.findById("wnd[1]/tbar[0]/btn[11]").press()

    time.sleep(25)
    excel = win32com.client.Dispatch('Excel.Application')
    time.sleep(10)

    excel.Quit()

    time.sleep(25)
    tabela_critica = pd.read_excel(r'C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels\critico.XLSX')
    tabela_estoque = pd.read_excel(r'C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels\estoque_critico.XLSX')


    critico_agrupe = tabela_critica.groupby(['Produto'],as_index=False)['Quantidade'].sum()
    estoque_agrupe = tabela_estoque.groupby(['Produto','Tipo de depósito'],as_index=False)['Quantidade'].sum()
    critico_cont = tabela_critica.groupby(['Produto'],as_index=False)['Quantidade'].count()

    data = {
    'Produto': [],
    'Estoque': [],
    'Roteiros': [],
    'virk': [],
    'Quantidade pedida': [],
    'Nº de Caixa':[],
    }
    copy = critico_agrupe['Produto']
    df = pd.DataFrame(data)
    df['Produto'] = copy
    df['Quantidade pedida'] = critico_agrupe['Quantidade']
    recebimento = {'9010' ,'8010', '8020','Y2ID','Y3ID','L0ID', 'VIRK'}
    







    for index, row in df.iterrows():

        valor_x = row['Produto']


        if valor_x in estoque_agrupe['Produto'].values:
            
            valor_W3PS = estoque_agrupe.loc[(estoque_agrupe['Produto'] == valor_x) & (estoque_agrupe['Tipo de depósito'] =='L2PS'), 'Quantidade'].values
            valor_virk = estoque_agrupe.loc[(estoque_agrupe['Produto'] == valor_x) & (estoque_agrupe['Tipo de depósito'] =='VIRK'), 'Quantidade'].values
            valor_recebimento = estoque_agrupe.loc[(estoque_agrupe['Produto'] == valor_x) & (estoque_agrupe['Tipo de depósito'].isin(recebimento)), 'Quantidade'].values
            df.loc[index, 'Estoque'] = valor_W3PS[0] if len(valor_W3PS) > 0 else 0
            df.loc[index, 'virk'] = valor_virk[0] if len(valor_virk) > 0 else 0
            df.loc[index, 'Roteiros'] = valor_recebimento[0] if len(valor_recebimento) > 0 else 0
            df.loc[index, 'Nº de Caixa'] = critico_cont.loc[critico_cont['Produto'] == valor_x, 'Quantidade'].values[0] if len(critico_cont.loc[critico_cont['Produto'] == valor_x, 'Quantidade'].values) > 0 else 0
        else:
            df.loc[index, 'Estoque'] = 0
            df.loc[index, 'virk'] = 0
            df.loc[index, 'Roteiros'] = 0
            df.loc[index, 'Nº de Caixa'] = 0

    wh = df.loc[(df['Estoque']>0), 'Produto'].values
    item_WH = pd.DataFrame({'Produto':wh})
    df.to_excel(r'C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels\Planilha_final.xlsx', index=False)
    item_WH.to_excel(r'C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels\Planilha_wh.xlsx', index=False)
    time.sleep(10)


    workbook = load_workbook(filename=r'C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels\Planilha_final.xlsx')


    worksheet = workbook.active


    alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

    worksheet.column_dimensions['A'].width = 20


    fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')


    for cell in worksheet[1]:
        cell.alignment = alignment
        cell.fill = fill


    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = alignment
    time.sleep(10)     
    workbook.save(filename=r'C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels\Planilha_final.xlsx')
    time.sleep(15)
    caminho_arquivo_excel = r'C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels\Planilha_wh.xlsx'

    df = pd.read_excel(caminho_arquivo_excel)
    nome_da_coluna = 'Produto'
    df[nome_da_coluna].to_clipboard(index=False)

    session.findById("wnd[0]").maximize
    session.findById("wnd[0]/tbar[0]/okcd").text = "/n/scwm/repl"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/usr/ctxtP_LGNUM").text = "ibi"
    session.findById("wnd[0]/usr/ctxtS_PSA-HIGH").setFocus
    session.findById("wnd[0]/usr/ctxtS_PSA-HIGH").caretPosition = 0
    session.findById("wnd[0]/usr/btn%_S_MATNR_%_APP_%-VALU_PUSH").press()
    session.findById("wnd[1]/tbar[0]/btn[24]").press()
    session.findById("wnd[1]/tbar[0]/btn[8]").press()
    session.findById("wnd[0]/tbar[1]/btn[8]").press()
    session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").setCurrentCell (-1,"")
    session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").selectAll()
    session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").pressToolbarButton ("EXEC")
    time.sleep(5)
    try:
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = "/n/scwm/mon"
        session.findById("wnd[0]").sendVKey (0)
        session.findById("wnd[0]/tbar[1]/btn[18]").press
        session.findById("wnd[0]/usr/shell/shellcont[0]/shell").expandNode ("C000000003")
        session.findById("wnd[0]/usr/shell/shellcont[0]/shell").selectedNode = "N000000032"
        session.findById("wnd[0]/usr/shell/shellcont[0]/shell").topNode = "C000000001"
        session.findById("wnd[0]/usr/shell/shellcont[0]/shell").doubleClickNode ("N000000032")
        session.findById("wnd[1]/usr/chkP_WOSTCA").selected = False
        session.findById("wnd[1]/usr/chkP_WOSTCO").selected = False
        session.findById("wnd[1]/usr/ctxtS_AAWHO-LOW").text = "k*"
        session.findById("wnd[1]/usr/ctxtS_QUEUE-LOW").text = "k*"
        session.findById("wnd[1]/usr/ctxtS_QUEUE-LOW").setFocus()
        session.findById("wnd[1]/usr/ctxtS_QUEUE-LOW").caretPosition = 2
        session.findById("wnd[1]/tbar[0]/btn[8]").press()
        session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").pressToolbarContextButton ("&MB_VARIANT")
        session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").selectContextMenuItem ("&LOAD")
        session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").setCurrentCell (1,"TEXT")
        session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").selectedRows = "1"
        session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").clickCurrentCell()
        session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").setCurrentCell (-1,"STATUS")
        session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").selectColumn ("STATUS")
        session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").pressToolbarButton( "&MB_FILTER")
        session.findById("wnd[1]/usr/ssub%_SUBSCREEN_FREESEL:SAPLSSEL:1105/ctxt%%DYN001-LOW").text = "b"
        session.findById("wnd[1]/usr/ssub%_SUBSCREEN_FREESEL:SAPLSSEL:1105/ctxt%%DYN001-LOW").caretPosition = 1
        session.findById("wnd[1]").sendVKey (0)
        session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").setCurrentCell (-1,"")
        session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").selectAll()
        session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").pressToolbarContextButton ("METHODS")
        session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").selectContextMenuItem ("@M00002")
        session.findById("wnd[1]/usr/sub:SAPLSPO4:0300/ctxtSVALD-VALUE[1,21]").text = "PATR"
        session.findById("wnd[1]/usr/sub:SAPLSPO4:0300/ctxtSVALD-VALUE[1,21]").caretPosition = 4
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        if session.ActiveWindow.Name == "wnd[1]":
                    
            session.findById("wnd[1]").close()
            
        session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").setCurrentCell (-1,"")
        session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").selectAll()
        session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").pressToolbarContextButton ("METHODS")
        session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").selectContextMenuItem ("@M00006")
        if session.ActiveWindow.Name == "wnd[1]":
                    
            session.findById("wnd[1]").close()
    except:
        pass
    time.sleep(5)
    session.findById("wnd[0]").maximize
    session.findById("wnd[0]/tbar[0]/okcd").text = "/n/scwm/mon"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/usr/shell/shellcont[0]/shell").expandNode ("C000000003")
    session.findById("wnd[0]/usr/shell/shellcont[0]/shell").selectedNode = "N000000032"
    session.findById("wnd[0]/usr/shell/shellcont[0]/shell").topNode = "C000000001"
    session.findById("wnd[0]/usr/shell/shellcont[0]/shell").doubleClickNode ("N000000032")
    session.findById("wnd[1]/usr/chkP_WOSTCO").selected = False
    session.findById("wnd[1]/usr/chkP_WOSTPR").selected = False
    session.findById("wnd[1]/usr/chkP_WOSTCA").selected = False
    session.findById("wnd[1]/usr/ctxtS_HDRWPT-LOW").text = "REP3"
    session.findById("wnd[1]/usr/ctxtS_HDRWPT-LOW").setFocus()
    session.findById("wnd[1]/usr/ctxtS_HDRWPT-LOW").caretPosition = 4
    session.findById("wnd[1]/tbar[0]/btn[8]").press()
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").pressToolbarContextButton ("&MB_VARIANT")
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").selectContextMenuItem ("&LOAD")
    session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").setCurrentCell (1,"TEXT")
    session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").selectedRows = "1"
    session.findById("wnd[1]/usr/subSUB_CONFIGURATION:SAPLSALV_CUL_LAYOUT_CHOOSE:0500/cntlD500_CONTAINER/shellcont/shell").clickCurrentCell()
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").setCurrentCell (-1,"CREATED_BY")
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").selectColumn ("CREATED_BY")
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").pressToolbarButton ("&MB_FILTER")
    session.findById("wnd[1]/usr/ssub%_SUBSCREEN_FREESEL:SAPLSSEL:1105/ctxt%%DYN001-LOW").text = "HE65465" #USUARIO DE QUEM VAI USAR
    session.findById("wnd[1]/usr/ssub%_SUBSCREEN_FREESEL:SAPLSSEL:1105/ctxt%%DYN001-LOW").caretPosition = 7
    session.findById("wnd[1]").sendVKey (0)
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").setCurrentCell (-1,"")
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").selectAll()
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").pressToolbarContextButton ("METHODS")
    session.findById("wnd[0]/usr/shell/shellcont[1]/shell/shellcont[0]/shell").selectContextMenuItem ("@WO0003")
    session.findById("wnd[1]/usr/ctxtP_QUEUE").text = "RAPIDOW3P3"
    session.findById("wnd[1]/usr/ctxtP_QUEUE").caretPosition = 10
    session.findById("wnd[1]/tbar[0]/btn[8]").press()
    if session.ActiveWindow.Name == "wnd[1]":
                
            session.findById("wnd[1]").close()
    time.sleep(10)
except:
    pass

    

tabela_critica = pd.read_excel(r'C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels\critico.XLSX')
tabela_estoque = pd.read_excel(r'C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels\estoque_critico.XLSX')


critico_agrupe = tabela_critica.groupby(['Produto'],as_index=False)['Quantidade'].sum()
estoque_agrupe = tabela_estoque.groupby(['Produto','Tipo de depósito'],as_index=False)['Quantidade'].sum()
critico_cont = tabela_critica.groupby(['Produto'],as_index=False)['Quantidade'].count()


df = pd.read_excel(r'C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels\Planilha_final.xlsx')


recebimentos = df.loc[(df['Roteiros']>0)&(df['Estoque']== 0 ), 'Produto'].values


n = len(recebimentos)
data = {
    'Produto': [None] * n,
    'Recebimento': [None] * n,
    'rota': [None] * n,
    'id': [None] * n
}
data['Produto'] = recebimentos

itens_recebimento = pd.DataFrame(data)

##
rota ={'ROT1', 'KART'}
rec ={'9010','8010','8020'}
ids ={'Y2ID','Y3ID'}
for index, row in itens_recebimento.iterrows():

    valor_x = row['Produto']
    if valor_x in estoque_agrupe['Produto'].values:
        valor_rec = estoque_agrupe.loc[(estoque_agrupe['Produto'] == valor_x) & (estoque_agrupe['Tipo de depósito'].isin(rec)), 'Quantidade'].values
        valor_rota = estoque_agrupe.loc[(estoque_agrupe['Produto'] == valor_x) & (estoque_agrupe['Tipo de depósito'].isin(rota)), 'Quantidade'].values
        valor_ids = estoque_agrupe.loc[(estoque_agrupe['Produto'] == valor_x) & (estoque_agrupe['Tipo de depósito'].isin(ids)), 'Quantidade'].values
        itens_recebimento.loc[index, 'Recebimento'] = valor_rec[0] if len(valor_rec) > 0 else 0
        itens_recebimento.loc[index, 'rota'] = valor_rota[0] if len(valor_rota) > 0 else 0
        itens_recebimento.loc[index, 'id'] = valor_ids[0] if len(valor_ids) > 0 else 0
    else:
        itens_recebimento.loc[index, 'Recebimento'] = 0
        itens_recebimento.loc[index, 'rota'] = 0
        itens_recebimento.loc[index, 'id'] = 0
rec_borda = itens_recebimento.style.set_properties(**{'border': '1px solid black'})
rec_borda.to_excel(r'C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels\Planilha_recebimento.xlsx', index=False)
time.sleep(8)

workbook = openpyxl.load_workbook(r'C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels\Planilha_recebimento.xlsx')
worksheet = workbook['Sheet1']

for row in worksheet.rows:
    for cell in row:
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
        cell.font = openpyxl.styles.Font(size=12)
        cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin', color='FF000000'),
                                            right=openpyxl.styles.Side(border_style='thin', color='FF000000'),
                                            top=openpyxl.styles.Side(border_style='thin', color='FF000000'),
                                            bottom=openpyxl.styles.Side(border_style='thin', color='FF000000'))
        worksheet.column_dimensions[cell.column_letter].auto_size = True

workbook.save(r'C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels\Planilha_recebimento.xlsx')

time.sleep(8)

#printer_name = '\\\\CANPMAN22'
#printer_name = '\\\\IBIPLOG01'  
#file_path = r'C:\Users\HE65465\AGCO Corp\Fiedler, Fabio - 01_ENG_Logistica\12_KPIs\Critico\excels\Planilha_recebimento.xlsx'
#Imprimir arquivo
#win32api.ShellExecute(
 #   0,
  #  'print',
 #   file_path,
 #   f'/d:"{printer_name}"',
 #   '.',
 #   0
#)
#time.sleep(2)

# Diretório onde os arquivos estão localizados
diretorio = "C:/Users/HE65465/AGCO Corp/Fiedler, Fabio - 01_ENG_Logistica/12_KPIs/Critico/excels/"

# Nomes dos arquivos
arquivo1 = "Planilha_recebimento.xlsx"
arquivo2 = "Planilha_wh.xlsx"

# Construindo os caminhos completos para os arquivos
file_path1 = os.path.join(diretorio, arquivo1)
file_path2 = os.path.join(diretorio, arquivo2)

# Nome da impressora
printer_name = '\\\\IBIPLOG01' 

# Imprimindo o primeiro arquivo
win32api.ShellExecute(0, 'print', file_path1, f'/d:"{printer_name}"', '.', 0)
time.sleep(3)

# Imprimindo o segundo arquivo
win32api.ShellExecute(0, 'print', file_path2, f'/d:"{printer_name}"', '.', 0)
time.sleep(3)